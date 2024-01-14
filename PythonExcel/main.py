from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook(r'C:\Users\kk\Downloads\Book1.xlsx')
ws = wb.active
print(ws)

for row in range(2,50):
    for col in range (1,4):
        char = get_column_letter(col)
        print(ws[char + str(row)])
        print(ws[char + str(row)].value)

# ws.save()

